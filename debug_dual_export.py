import sys
import os
import pandas as pd
from datetime import datetime

# Agregar Actividades al path
sys.path.append(os.path.join(os.getcwd(), "Actividades"))

from config import TEMPLATE_EXCEL, TEMPLATE_INFORME_FINAL
from export_service import generar_informe_template
from export_final_service import generar_informe_final_resumen

# Data de prueba
data = {
    'USUARIO': ['test', 'test'],
    'TIPO DE ACTIVIDAD': ['Actividad A', 'Actividad B'],
    'FECHA': [datetime.now(), datetime.now()],
    'DEPENDENCIA': ['Dep 1', 'Dep 2'],
    'SOLICITANTE': ['Sol 1', 'Sol 2'],
    'TIPO DE SOLICITUD': ['Tipo 1', 'Tipo 2'],
    'MEDIO DE SOLICITUD': ['Medio 1', 'Medio 2'],
    'CUMPLIDO': ['Sí', 'Sí'],
    'FECHA ATENCIÓN': [datetime.now().strftime('%Y-%m-%d')] * 2,
    'OBSERVACIONES': ['Obs 1', 'Obs 2']
}
df = pd.DataFrame(data)
contrato_data = {'nro': '123', 'objeto': 'Prueba', 'nombre': 'Tester', 'supervisor': 'Boss'}

print(f"--- VALIDACIÓN DUAL ---")
print(f"Plantilla Detallada: {TEMPLATE_EXCEL}")
print(f"Plantilla Final: {TEMPLATE_INFORME_FINAL}")

# 1. Generar Detallado
path_detallado = "debug_detallado.xlsx"
generar_informe_template(df, path_detallado, contrato_data=contrato_data)
print(f"Reporte Detallado generado en {path_detallado}")

# 2. Generar Final
path_final = "debug_final.xlsx"
generar_informe_final_resumen(df, path_final, contrato_data=contrato_data)
print(f"Reporte Final generado en {path_final}")

# 3. Verificación de cabeceras en los generados
import openpyxl
wb_d = openpyxl.load_workbook(path_detallado)
ws_d = wb_d.active
print(f"Cabecera Detallado (A7): {ws_d.cell(7,1).value}")

wb_f = openpyxl.load_workbook(path_final)
ws_f = wb_f.active
print(f"Cabecera Final (A7): {ws_f.cell(7,1).value}")
