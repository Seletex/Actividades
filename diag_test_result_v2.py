
import openpyxl
import os

path = "test_informe_final.xlsx"
if os.path.exists(path):
    print(f"Inspeccionando resultado: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    print("\nDetalle de filas clave:")
    for row in range(7, 22):
        col1 = ws.cell(row=row, column=1).value or ""
        col2 = ws.cell(row=row, column=2).value or ""
        col8 = ws.cell(row=row, column=8).value or ""
        
        # Detectar si hay merge en col 1
        is_merged_c1 = any(row >= r.min_row and row <= r.max_row and 1 >= r.min_col and 1 <= r.max_col for r in ws.merged_cells.ranges)
        # Detectar si hay merge en col 8
        is_merged_c8 = any(row >= r.min_row and row <= r.max_row and 8 >= r.min_col and 8 <= r.max_col for r in ws.merged_cells.ranges)
        
        m1 = "[M]" if is_merged_c1 else "[ ]"
        m8 = "[M]" if is_merged_c8 else "[ ]"
        
        print(f"Fila {row:02d} | C1: {m1} '{col1:<30}' | C2: '{col2:<20}' | C8: {m8} '{col8}'")
else:
    print(f"No se encontró el archivo {path}")
