"""
Servicio de exportación y generación de reportes.
Separado de database.py para responsabilidad única.
"""

import os
import copy
import pandas as pd
from datetime import datetime
from config import TEMPLATE_EXCEL, logger
from database import cargar_registros
from utils import medir_tiempo


@medir_tiempo
def exportar_registros_filtrados(fecha_inicio=None, fecha_fin=None, usuario=None, actividad=None):
    """Exporta registros filtrados. Retorna (DataFrame, dict_estadísticas)"""
    try:
        df = cargar_registros(usuario)
        if df.empty:
            return pd.DataFrame(), {}
        
        # Parsear fechas
        if 'FECHA' in df.columns:
            df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        
        # Aplicar filtros
        if fecha_inicio:
            df = df[df['FECHA'] >= pd.to_datetime(fecha_inicio)]
        if fecha_fin:
            df = df[df['FECHA'] <= pd.to_datetime(fecha_fin)]
        if actividad and actividad != 'Todas' and 'TIPO DE ACTIVIDAD' in df.columns:
            df = df[df['TIPO DE ACTIVIDAD'] == actividad]
        
        # Agrupamiento y ordenamiento solicitado
        if not df.empty:
            sort_cols = []
            if 'TIPO DE ACTIVIDAD' in df.columns: sort_cols.append('TIPO DE ACTIVIDAD')
            if 'FECHA' in df.columns: sort_cols.append('FECHA')
            if sort_cols:
                df = df.sort_values(by=sort_cols)
        
        stats = _calcular_estadisticas(df)
        return df, stats
    except Exception as e:
        logger.error(f"Error exportando registros: {e}")
        return pd.DataFrame(), {}


def _calcular_estadisticas(df):
    """Calcula estadísticas básicas de un DataFrame"""
    if df.empty:
        return {}
    
    stats = {
        'total_registros': len(df),
        'fecha_inicio': _format_fecha(df, 'min'),
        'fecha_fin': _format_fecha(df, 'max'),
    }
    
    # Conteos por columna
    for col, key in [('TIPO DE ACTIVIDAD', 'conteo_por_actividad'),
                     ('TIPO DE SOLICITUD', 'conteo_por_solicitud'),
                     ('MEDIO DE SOLICITUD', 'conteo_por_medio')]:
        if col in df.columns:
            counts = df[col].value_counts().to_dict()
            stats[key] = {str(k): int(v) for k, v in counts.items()}
    
    return stats


def _format_fecha(df, func):
    """Formatea fecha min/max de un DataFrame"""
    if 'FECHA' not in df.columns or df['FECHA'].empty:
        return 'N/A'
    try:
        val = getattr(df['FECHA'], func)()
        return val.strftime('%Y-%m-%d') if pd.notna(val) else 'N/A'
    except Exception:
        return 'N/A'


@medir_tiempo
def obtener_estadisticas_exportacion(usuario=None, fecha_inicio=None, fecha_fin=None):
    """Obtiene estadísticas generales y datos para gráficos con soporte de filtros"""
    empty_result = {
        'fecha_min': 'N/A', 'fecha_max': 'N/A',
        'total_registros': 0, 'total_tipos_actividad': 0,
        'ultima_exportacion': 'Nunca',
        'chart_actividades': {'labels': [], 'data': []},
        'chart_cumplimiento': {'labels': [], 'data': []},
        'chart_linea': {'labels': [], 'data': []},
        'usuarios': []
    }
    
    try:
        df, _ = exportar_registros_filtrados(
            fecha_inicio=fecha_inicio, 
            fecha_fin=fecha_fin, 
            usuario=usuario
        )
        if df.empty:
            return empty_result
        
        # Parseo de fechas (exportar_registros_filtrados ya hace parte del proceso)
        if 'FECHA' in df.columns:
            df['FECHA_DT'] = pd.to_datetime(df['FECHA'], errors='coerce')
            df = df.dropna(subset=['FECHA_DT'])
        
        if df.empty:
            return empty_result
        
        # Gráfico: Actividades (Todas, según petición del usuario)
        counts_act = df['TIPO DE ACTIVIDAD'].value_counts()
        chart_actividades = {
            'labels': counts_act.index.tolist(),
            'data': counts_act.values.tolist()
        }
        
        # Gráfico: Cumplimiento
        cumplimiento = df['CUMPLIDO'].value_counts() if 'CUMPLIDO' in df.columns else pd.Series()
        chart_cumplimiento = {
            'labels': cumplimiento.index.tolist(),
            'data': cumplimiento.values.tolist()
        }
        
        # Gráfico: Línea temporal
        df_sorted = df.sort_values('FECHA_DT')
        linea = df_sorted['FECHA_DT'].dt.date.value_counts().sort_index()
        # Si hay demasiados días, mostrar los últimos 90 para no saturar
        if len(linea) > 90:
            linea = linea.tail(90)
            
        chart_linea = {
            'labels': [d.strftime('%d/%m') for d in linea.index],
            'data': linea.values.tolist()
        }
        
        # Estadística por usuario
        user_stats = []
        if 'USUARIO' in df.columns:
            for user, group in df.groupby('USUARIO'):
                total_user = len(group)
                cumplidos = len(group[group['CUMPLIDO'] == 'Sí'])
                porcentaje = f"{(cumplidos/total_user)*100:.1f}%" if total_user > 0 else "0%"
                ultima = group['FECHA_DT'].max().strftime('%Y-%m-%d %H:%M') if not group.empty else "N/A"
                user_stats.append({
                    'usuario': user,
                    'total': total_user,
                    'cumplimiento': porcentaje,
                    'ultima': ultima
                })
        
        return {
            'fecha_min': df['FECHA_DT'].min().strftime('%Y-%m-%d'),
            'fecha_max': df['FECHA_DT'].max().strftime('%Y-%m-%d'),
            'total_registros': len(df),
            'total_tipos_actividad': df['TIPO DE ACTIVIDAD'].nunique() if 'TIPO DE ACTIVIDAD' in df.columns else 0,
            'ultima_exportacion': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'chart_actividades': chart_actividades,
            'chart_cumplimiento': chart_cumplimiento,
            'chart_linea': chart_linea,
            'usuarios': user_stats
        }
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {e}")
        return empty_result


@medir_tiempo
def generar_reporte_excel(df, estadisticas, output_path):
    """Genera un archivo Excel con datos + estadísticas"""
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Registros', index=False)
            
            stats_df = pd.DataFrame([
                ['Total de registros', estadisticas.get('total_registros', 0)],
                ['Fecha inicio', estadisticas.get('fecha_inicio', 'N/A')],
                ['Fecha fin', estadisticas.get('fecha_fin', 'N/A')]
            ], columns=['Métrica', 'Valor'])
            stats_df.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            if 'conteo_por_actividad' in estadisticas:
                act_df = pd.DataFrame(
                    estadisticas['conteo_por_actividad'].items(),
                    columns=['Actividad', 'Cantidad']
                )
                act_df.to_excel(writer, sheet_name='Estadísticas', startrow=5, index=False)
        
        return True
    except Exception as e:
        logger.error(f"Error generando reporte Excel: {e}")
        return False


@medir_tiempo
def generar_informe_template(df, output_path, contrato_data=None):
    """Genera informe usando la plantilla Excel institucional"""
    try:
        import openpyxl
        from config import TEMPLATE_EXCEL
        logger.info(f"DEBUG: generar_informe_template usando plantilla IMPORTADA: {TEMPLATE_EXCEL}")
        if not os.path.exists(TEMPLATE_EXCEL):
            logger.error(f"Plantilla no encontrada: {TEMPLATE_EXCEL}")
            return False
            
        try:
            wb = openpyxl.load_workbook(TEMPLATE_EXCEL)
            ws = wb.active
            logger.info("Plantilla cargada correctamente")
        except Exception as e:
            logger.error(f"Error al cargar el libro de Excel: {e}")
            return False
        
        # Encabezados: Datos del Contrato (Filas 2-5, Columna 3)
        if contrato_data:
            if contrato_data.get('nro'):
                ws.cell(row=2, column=3, value=contrato_data['nro'].upper())
            if contrato_data.get('objeto'):
                ws.cell(row=3, column=3, value=contrato_data['objeto'].upper())
            
            # El nombre se pone en la fila 4 por defecto si no hay contrato_data['nombre']
            nombre_contratista = contrato_data.get('nombre', '').upper()
            if not nombre_contratista:
                usuarios = df['USUARIO'].unique() if 'USUARIO' in df.columns else []
                nombre_contratista = usuarios[0].upper() if len(usuarios) == 1 else "VARIOS"
            
            ws.cell(row=4, column=3, value=nombre_contratista)
            
            if contrato_data.get('cedula'):
                ws.cell(row=5, column=3, value=contrato_data['cedula'])

        # Rango de fechas (Fila 6)
        if not df.empty and 'FECHA' in df.columns:
            fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
            if not fechas_dt.empty:
                ws.cell(row=6, column=3,
                        value=f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}")
        
        # Guardar estilos base de fila 8
        base_styles = []
        for col in range(1, 10):
            cell = ws.cell(row=8, column=col)
            base_styles.append({
                'font': copy.copy(cell.font),
                'border': copy.copy(cell.border),
                'fill': copy.copy(cell.fill),
                'alignment': copy.copy(cell.alignment),
                'number_format': cell.number_format
            })
        
        # Diccionario de meses en español
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        ahora = datetime.now()
        mes_actual = meses[ahora.month - 1]
        anio_actual = ahora.year

        # Limpiar filas 8 en adelante (hasta 200 para seguridad)
        for row_idx in range(8, 200):
            for col_idx in range(1, 11):
                ws.cell(row=row_idx, column=col_idx, value=None)
        
        # Desenlazar celdas desde la fila 8 en adelante
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.min_row >= 8:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass
        
        # Escribir datos con cortes por actividad
        current_row = 8
        ultima_actividad = None
        conteo_actividad = 0
        total_general = 0
        
        # Asegurar ordenamiento para el reporte
        df_reporte = df.copy()
        if 'TIPO DE ACTIVIDAD' in df_reporte.columns:
            df_reporte = df_reporte.sort_values(by=['TIPO DE ACTIVIDAD', 'FECHA'])

        for _, row in df_reporte.iterrows():
            actividad_actual = row.get('TIPO DE ACTIVIDAD', '')
            
            # Si cambia la actividad, mostrar subtotal de la anterior
            if ultima_actividad is not None and actividad_actual != ultima_actividad:
                ws.cell(row=current_row, column=1, value="ACTIVIDADES: ")
                # Fusionar celdas de columna 2 a 8 para el conteo
                ws.cell(row=current_row, column=2, value=conteo_actividad).font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                
                for c in range(1, 10):
                    ws.cell(row=current_row, column=c).border = copy.copy(base_styles[c-1]['border'])
                current_row += 1
                conteo_actividad = 0
            
            fecha = row.get('FECHA', '')
            if hasattr(fecha, 'strftime'):
                fecha = fecha.strftime('%Y-%m-%d')
            
            fecha_atencion = row.get('FECHA ATENCIÓN', '')
            if hasattr(fecha_atencion, 'strftime'):
                fecha_atencion = fecha_atencion.strftime('%Y-%m-%d')
            
            valores = [
                actividad_actual, fecha,
                row.get('DEPENDENCIA', ''), row.get('SOLICITANTE', ''),
                row.get('TIPO DE SOLICITUD', ''), row.get('MEDIO DE SOLICITUD', ''),
                row.get('CUMPLIDO', 'Sí'), fecha_atencion,
                row.get('OBSERVACIONES', '')
            ]
            
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=valor)
                if col_idx <= len(base_styles):
                    style = base_styles[col_idx - 1]
                    cell.font = copy.copy(style['font'])
                    cell.border = copy.copy(style['border'])
                    cell.fill = copy.copy(style['fill'])
                    cell.alignment = copy.copy(style['alignment'])
                    cell.number_format = style['number_format']
            
            ultima_actividad = actividad_actual
            conteo_actividad += 1
            total_general += 1
            current_row += 1
            
            if current_row > 2000: break # Limite de seguridad
            
        # Último subtotal y Gran Total
        if ultima_actividad:
            ws.cell(row=current_row, column=1, value="ACTIVIDADES: ")
            ws.cell(row=current_row, column=2, value=conteo_actividad).font = openpyxl.styles.Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            
            for c in range(1, 10):
                ws.cell(row=current_row, column=c).border = copy.copy(base_styles[c-1]['border'])

            current_row += 1
            
            ws.cell(row=current_row, column=1, value="TOTAL GENERAL").font = openpyxl.styles.Font(bold=True, size=11)
            ws.cell(row=current_row, column=2, value=total_general).font = openpyxl.styles.Font(bold=True, size=11)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
            ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            
            for c in range(1, 10):
                ws.cell(row=current_row, column=c).border = copy.copy(base_styles[c-1]['border'])

            current_row += 1
        # Fecha de informe
        ws.cell(row=current_row, column=1, value="Fecha de informe:").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"{mes_actual} de {anio_actual}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
        ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        current_row += 1
        # Sección de firmas / Datos finales

        if contrato_data:
            if contrato_data.get('nombre'):
                ws.cell(row=current_row, column=1, value="Elaborado por:").font = openpyxl.styles.Font(bold=True)
                ws.cell(row=current_row, column=2, value=contrato_data['nombre'].upper())
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                current_row += 1
                ws.cell(row=current_row, column=1, value="CONTRATISTA:").font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                current_row += 1                
            if contrato_data.get('supervisor'):
                ws.cell(row=current_row, column=1, value="Vo.Bo:").font = openpyxl.styles.Font(bold=True)
                ws.cell(row=current_row, column=2, value=contrato_data['supervisor'].upper())
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                current_row += 1
                ws.cell(row=current_row, column=1, value="SUPERVISOR:").font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                current_row += 1
            
            current_row += 1


        try:
            wb.save(output_path)
            logger.info(f"Informe guardado exitosamente en: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error al guardar el archivo de salida {output_path}: {e}")
            return False
    except Exception as e:
        logger.exception(f"Excepción no controlada en generar_informe_template: {e}")
        return False
