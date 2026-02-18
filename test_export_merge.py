
import sys
import os
import pandas as pd
from datetime import datetime
import openpyxl

# Agregar el directorio actual al path
sys.path.append(os.getcwd())

try:
    from config import TEMPLATE_EXCEL, logger
    from export_service import generar_informe_template
    
    # Simular registros con dos actividades diferentes para ver subtotales
    data = {
        'ID': [1, 2, 3],
        'TIPO DE ACTIVIDAD': ['Soporte Técnico', 'Soporte Técnico', 'Mantenimiento'],
        'FECHA': [datetime.now()] * 3,
        'DEPENDENCIA': ['Secretaría'] * 3,
        'SOLICITANTE': ['Juan Perez'] * 3,
        'TIPO DE SOLICITUD': ['Correctivo'] * 3,
        'MEDIO DE SOLICITUD': ['Teléfono'] * 3,
        'DESCRIPCIÓN': ['Prueba de fusión'] * 3,
        'CUMPLIDO': ['Sí'] * 3,
        'FECHA ATENCIÓN': [datetime.now().strftime('%Y-%m-%d')] * 3,
        'OBSERVACIONES': ['Sin observaciones'] * 3
    }
    df = pd.DataFrame(data)
    
    output_test = "test_merged_output.xlsx"
    print(f"Generando informe de prueba en: {output_test}")
    
    success = generar_informe_template(df, output_test)
    
    if success and os.path.exists(output_test):
        wb = openpyxl.load_workbook(output_test)
        ws = wb.active
        
        print("\nVerificando fusiones de celdas (B-H) en filas de totales:")
        ranges = ws.merged_cells.ranges
        
        rows_found = 0
        for r in sorted(list(ranges), key=lambda x: x.min_row):
            if r.min_col == 2 and r.max_col == 8:
                cell_val = ws.cell(row=r.min_row, column=1).value
                count_val = ws.cell(row=r.min_row, column=2).value
                print(f" - Fila {r.min_row:02d}: {r} | C1: '{cell_val}' | C2: '{count_val}'")
                rows_found += 1
        
        if rows_found >= 3: # 2 subtotales + 1 total general
            print("\n✅ Verificación exitosa: Se encontraron las fusiones esperadas.")
        else:
            print(f"\n⚠️ Advertencia: Solo se encontraron {rows_found} fusiones de tipo B-H.")
        
        # Opcional: limpiar
        # os.remove(output_test)
    else:
        print("❌ Error: No se pudo generar o encontrar el archivo de salida.")
            
except Exception as e:
    import traceback
    traceback.print_exc()
