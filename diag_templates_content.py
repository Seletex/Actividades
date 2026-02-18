import openpyxl
import os

files = [
    r"c:\Users\apoyosistemas\Documents\Python - copia\INFORME DE ACTIVIDADES - copia.xlsx",
    r"c:\Users\apoyosistemas\Documents\Python - copia\InformeFinal.XLSX"
]

for f in files:
    if os.path.exists(f):
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            print(f"Archivo: {os.path.basename(f)}")
            print(f"  Celda A1: {ws.cell(1,1).value}")
            print(f"  Celda B1: {ws.cell(1,2).value}")
            print(f"  Celda C2: {ws.cell(2,3).value}")
        except Exception as e:
            print(f"Error leyendo {f}: {e}")
    else:
        print(f"No existe: {f}")
