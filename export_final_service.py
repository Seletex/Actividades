import os
import pandas as pd
import openpyxl
import copy
from datetime import datetime
from config import TEMPLATE_INFORME_FINAL, logger
from utils import medir_tiempo

@medir_tiempo
def generar_informe_final_resumen(df, output_path, contrato_data=None):
    """
    Genera un informe concentrado contando actividades por tipo.
    Mantiene los encabezados institucionales de la plantilla.
    """
    try:
        logger.info(f"Generando Informe Final Concentrado. Registros: {len(df)}")
        if not os.path.exists(TEMPLATE_INFORME_FINAL):
            logger.error(f"Plantilla no encontrada: {TEMPLATE_INFORME_FINAL}")
            return False

        wb = openpyxl.load_workbook(TEMPLATE_INFORME_FINAL)
        ws = wb.active
        
        # 1. Agrupar y contar actividades
        if 'TIPO DE ACTIVIDAD' in df.columns:
            resumen = df['TIPO DE ACTIVIDAD'].value_counts().reset_index()
            resumen.columns = ['Actividad', 'Cantidad']
        else:
            resumen = pd.DataFrame(columns=['Actividad', 'Cantidad'])

        # 2. Encabezados Institucionales (misma lógica que el detallado)
        if contrato_data:
            if contrato_data.get('nro'):
                ws.cell(row=2, column=3, value=contrato_data['nro'].upper())
            if contrato_data.get('objeto'):
                ws.cell(row=3, column=3, value=contrato_data['objeto'].upper())
            
            nombre_contratista = contrato_data.get('nombre', '').upper()
            if not nombre_contratista:
                usuarios = df['USUARIO'].unique() if 'USUARIO' in df.columns else []
                nombre_contratista = usuarios[0].upper() if len(usuarios) == 1 else "VARIOS"
            ws.cell(row=4, column=3, value=nombre_contratista)
            
            if contrato_data.get('cedula'):
                ws.cell(row=5, column=3, value=contrato_data['cedula'])

        # Rango de fechas (Fila 6)
        if not df.empty and 'FECHA' in df.columns:
            fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
            if not fechas_dt.empty:
                ws.cell(row=6, column=3,
                        value=f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}")

        # 3. Guardar estilos base de fila 7 (Cabecera) y fila 8 (Datos)
        # En la plantilla V3, la fila 7 tiene "TIPO DE ACTIVIDAD" y "CANTIDAD"
        header_styles = []
        for col in range(1, 11):
            cell = ws.cell(row=7, column=col)
            header_styles.append({
                'font': copy.copy(cell.font),
                'border': copy.copy(cell.border),
                'fill': copy.copy(cell.fill),
                'alignment': copy.copy(cell.alignment)
            })

        # 4. Limpieza de área de datos (Filas 7 en adelante)
        # Desenlazar celdas PRIMERO
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.min_row >= 7:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception: pass

        # Luego limpiar para reconstruir con los datos actuales
        for row in ws.iter_rows(min_row=7, max_row=300, min_col=1, max_col=11):
            for cell in row:
                cell.value = None

        # 5. Escribir Cabecera de Tabla (Fila 7)
        current_row = 7
        ws.cell(row=current_row, column=1, value="TIPO DE ACTIVIDAD")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=8, value="CANTIDAD")
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=10)
        
        # Aplicar estilos de cabecera
        for col_idx in range(1, 11):
            cell = ws.cell(row=current_row, column=col_idx)
            style = header_styles[col_idx-1]
            cell.font = style['font']
            cell.border = style['border']
            cell.fill = style['fill']
            cell.alignment = style['alignment']
        
        current_row += 1
        
        # 6. Escribir Datos de Resumen
        for _, r in resumen.iterrows():
            # Actividad (A-G)
            ws.cell(row=current_row, column=1, value=r['Actividad'])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            # Cantidad (H-J)
            ws.cell(row=current_row, column=8, value=r['Cantidad'])
            ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=10)
            
            # Bordes y alineación para la fila de datos
            for c in range(1, 11):
                cell = ws.cell(row=current_row, column=c)
                cell.border = header_styles[c-1]['border']
                cell.alignment = header_styles[c-1]['alignment']
                # Si es la columna de cantidad (H), centrar
                if c >= 8:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
        # Total General
        ws.cell(row=current_row, column=1, value="TOTAL GENERAL DE ACTIVIDADES:").font = openpyxl.styles.Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=8, value=resumen['Cantidad'].sum()).font = openpyxl.styles.Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=10)
        
        for c in range(1, 11):
            cell = ws.cell(row=current_row, column=c)
            cell.border = header_styles[c-1]['border']
            cell.alignment = header_styles[c-1]['alignment']
            if c >= 8:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        current_row += 1
        
        # 7. Fecha e Informe (mismo estilo que el detallado)
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        ahora = datetime.now()
        ws.cell(row=current_row, column=1, value="Fecha de informe:").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"{meses[ahora.month-1]} de {ahora.year}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
        current_row += 1
        
        # Firmas
        if contrato_data:
            start_firmas = current_row
            # Nombre Contratista
            if contrato_data.get('nombre'):
                ws.cell(row=current_row, column=1, value="Elaborado por:").font = openpyxl.styles.Font(bold=True)
                ws.cell(row=current_row, column=2, value=contrato_data['nombre'].upper())
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                current_row += 1
                ws.cell(row=current_row, column=1, value="CONTRATISTA:").font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                current_row += 1
            
            # Supervisor
            if contrato_data.get('supervisor'):
                ws.cell(row=current_row, column=1, value="Vo.Bo:").font = openpyxl.styles.Font(bold=True)
                ws.cell(row=current_row, column=2, value=contrato_data['supervisor'].upper())
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                current_row += 1
                ws.cell(row=current_row, column=1, value="SUPERVISOR:").font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                current_row += 1

        wb.save(output_path)
        return True
    except Exception as e:
        logger.error(f"Error generando informe final: {e}")
        return False
